import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class ExcelOperations:

    @staticmethod
    def csv_to_excel(csv_file_path, excel_file_path):
        wb = Workbook()
        ws = wb.active

        with open(csv_file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)

        wb.save(excel_file_path)

    

    @staticmethod
    def format_mps_overview_excel(file_path):
        wb = load_workbook(file_path)
        ws = wb.active

        # Title row formatting: bold with a thick bottom border
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.border = Border(bottom=Side(style='thick'))
            cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right")

        # Data rows formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column != 1:  # Apply number formatting for non-first columns
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "right")

                # Apply thick bottom border for the second-last and last row
                if row[0].row in [ws.max_row, ws.max_row - 1]:
                    cell.border = Border(bottom=Side(style='thick'))

                # Bold for the 'Total Logs' in the last row
                if row[0].row == ws.max_row and cell.column == 1:
                    cell.font = Font(bold=True)

        # No borders except for specified rows
        # Adjust column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = max_length + 2

        # Freeze title row
        ws.freeze_panes = "A2"

        # Save the workbook
        wb.save(file_path)




# Usage example (Uncomment and adjust the file paths to test)
#ExcelOperations.csv_to_excel('Z:/Final Project/Data_export_tests/log_source_overview.csv', 'Z:/Final Project/Data_export_tests/log_source_overview.xlsx')
#ExcelOperations.format_excel('Z:/Final Project/Data_export_tests/log_source_overview.xlsx')
